#!/usr/bin/env python3
"""Concilia un conteo físico (Excel) contra el inventario real de Shopify.

Ver INSTRUCTIVO-CONCILIAR-INVENTARIO.md para el proceso completo. Cruza
por SKU (`No Parte` en el Excel contra `variant.sku` en Shopify) y, para
las filas sin SKU, intenta un segundo cruce por nombre (`Descripción`
contra el título del producto). Sube a Shopify lo que cambió, y regresa
el mismo Excel coloreado con el resultado.

El cruce por nombre es deliberadamente conservador: solo actualiza
automático cuando hay una sola coincidencia confiable (exacta tras
normalizar, o aproximada por encima del 90% sin empate con otra). Si hay
ambigüedad, la fila queda gris con los candidatos anotados — nunca
adivina entre dos productos parecidos.

Uso:
    SHOPIFY_ADMIN_TOKEN=shpat_... python3 scripts/conciliar-inventario.py entrada.xlsx salida.xlsx

Variables de entorno:
    SHOPIFY_ADMIN_TOKEN  (obligatoria) token Admin API con scopes de inventario
    SHOPIFY_STORE        dominio myshopify (default: wfuxvx-yn.myshopify.com)
"""

import copy
import difflib
import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request

import openpyxl
from openpyxl.styles import PatternFill

API_VERSION = "2024-01"

COLOR_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
COLOR_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
COLOR_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
COLOR_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def api_get(path, token, store, params=None):
    url = f"https://{store}/admin/api/{API_VERSION}{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url)
    req.add_header("X-Shopify-Access-Token", token)
    with urllib.request.urlopen(req) as resp:
        link = resp.headers.get("Link", "")
        return json.loads(resp.read()), link


def next_page_url(link_header):
    if not link_header:
        return None
    for part in link_header.split(","):
        if 'rel="next"' in part:
            return part.split(";")[0].strip().strip("<>")
    return None


def fetch_all_products(token, store):
    products = []
    url = f"https://{store}/admin/api/{API_VERSION}/products.json?limit=250"
    while url:
        req = urllib.request.Request(url)
        req.add_header("X-Shopify-Access-Token", token)
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
            products.extend(data.get("products", []))
            url = next_page_url(resp.headers.get("Link", ""))
        time.sleep(0.5)
    return products


def build_sku_map(products):
    sku_map = {}
    for product in products:
        for variant in product.get("variants", []):
            sku = (variant.get("sku") or "").strip()
            if not sku:
                continue
            sku_map[sku] = {
                "product_title": product["title"],
                "variant_id": variant["id"],
                "inventory_item_id": variant["inventory_item_id"],
                "available": variant.get("inventory_quantity", 0),
            }
    return sku_map


def build_barcode_map(products):
    """Indexa por `barcode`, donde vive el "Codigo B1" del POS.

    Es la llave que cierra los huecos del cruce por SKU: filas del conteo
    sin `No Parte`, y productos cuyo SKU en Shopify viene del catalogo del
    fabricante en vez del codigo interno. Ver scripts/vincular-codigo-b1.py
    e INSTRUCTIVO-CONCILIAR-INVENTARIO.md.
    """
    barcode_map = {}
    for product in products:
        for variant in product.get("variants", []):
            barcode = (variant.get("barcode") or "").strip()
            if not barcode:
                continue
            barcode_map[barcode] = {
                "product_title": product["title"],
                "variant_id": variant["id"],
                "inventory_item_id": variant["inventory_item_id"],
                "available": variant.get("inventory_quantity", 0),
            }
    return barcode_map


def normalize_name(text):
    """Uppercase, sin acentos, sin puntuación, espacios colapsados.

    Suficiente para que "Caña de Pescar Shimano Sellus Spinning 5'8\""
    y "CAÑA DE PESCAR SHIMANO SELLUS SPINNING 5'8" (POS en mayúsculas,
    Shopify con mayúsculas/minúsculas normales) se reconozcan como el
    mismo texto.
    """
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", str(text))
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def build_title_index(products):
    """Normalized product title -> info de la variante (asume 1 variante
    por producto, cierto para este catálogo — build_sku_map ya lo
    confirma: mismo número de productos que de variantes con SKU)."""
    index = {}
    for product in products:
        variants = product.get("variants", [])
        if len(variants) != 1:
            continue  # ambiguo a qué variante actualizar, no se intenta por nombre
        variant = variants[0]
        norm = normalize_name(product["title"])
        if not norm:
            continue
        index.setdefault(norm, []).append(
            {
                "product_title": product["title"],
                "variant_id": variant["id"],
                "inventory_item_id": variant["inventory_item_id"],
                "available": variant.get("inventory_quantity", 0),
            }
        )
    return index


def find_by_name(descripcion, title_index):
    """Busca una descripción del conteo físico contra los títulos de Shopify.

    Devuelve (info, nota) si hay una sola coincidencia confiable, o
    (None, nota_explicando_por_que_no) en cualquier caso ambiguo — nunca
    elige entre dos candidatos parecidos.
    """
    norm = normalize_name(descripcion)
    if not norm:
        return None, "Sin código de parte ni descripción utilizable"

    exactas = title_index.get(norm)
    if exactas and len(exactas) == 1:
        return exactas[0], f"Vinculado por nombre exacto: \"{exactas[0]['product_title']}\""
    if exactas and len(exactas) > 1:
        return None, f"Sin código de parte; nombre coincide con {len(exactas)} productos distintos en Shopify — revisar a mano"

    titulos = list(title_index.keys())
    cercanos = difflib.get_close_matches(norm, titulos, n=3, cutoff=0.90)
    if len(cercanos) == 1:
        info = title_index[cercanos[0]][0]
        return info, f"Vinculado por nombre aproximado: \"{info['product_title']}\""
    if len(cercanos) > 1:
        candidatos = "; ".join(title_index[c][0]["product_title"] for c in cercanos)
        return None, f"Sin código de parte; {len(cercanos)} nombres parecidos en Shopify ({candidatos}) — revisar a mano"

    return None, "Sin código de parte; sin coincidencia por nombre en Shopify"


def set_inventory(token, store, location_id, inventory_item_id, available):
    url = f"https://{store}/admin/api/{API_VERSION}/inventory_levels/set.json"
    body = json.dumps(
        {"location_id": location_id, "inventory_item_id": inventory_item_id, "available": available}
    ).encode()
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("X-Shopify-Access-Token", token)
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        return {"error": e.read().decode()}


def main():
    if len(sys.argv) != 3:
        print("Uso: conciliar-inventario.py entrada.xlsx salida.xlsx", file=sys.stderr)
        sys.exit(1)

    input_path, output_path = sys.argv[1], sys.argv[2]
    token = os.environ.get("SHOPIFY_ADMIN_TOKEN")
    store = os.environ.get("SHOPIFY_STORE", "wfuxvx-yn.myshopify.com")
    if not token:
        print("Falta SHOPIFY_ADMIN_TOKEN", file=sys.stderr)
        sys.exit(1)

    print("Descargando productos de Shopify...")
    products = fetch_all_products(token, store)
    sku_map = build_sku_map(products)
    barcode_map = build_barcode_map(products)
    title_index = build_title_index(products)
    print(
        f"  {len(products)} productos, {len(sku_map)} con SKU, "
        f"{len(barcode_map)} con codigo B1 (barcode)"
    )

    shop_data, _ = api_get("/shop.json", token, store)
    location_id = shop_data["shop"]["primary_location_id"]

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}
    if "No Parte" not in col or "Existencia" not in col:
        print("El Excel debe tener columnas 'No Parte' y 'Existencia'", file=sys.stderr)
        sys.exit(1)

    col_antes = len(headers)
    col_estatus = len(headers) + 1
    col_nota = len(headers) + 2
    ws.cell(row=1, column=col_antes + 1, value="Existencia Shopify (antes)")
    ws.cell(row=1, column=col_estatus + 1, value="Estatus")
    ws.cell(row=1, column=col_nota + 1, value="Nota")

    sku_counts = {}
    rows = list(ws.iter_rows(min_row=2))
    for row in rows:
        sku = row[col["No Parte"]].value
        if sku:
            sku = str(sku).strip()
            sku_counts[sku] = sku_counts.get(sku, 0) + 1

    counts = {"verde": 0, "amarillo": 0, "rojo": 0, "gris": 0}
    vinculados_por_nombre = 0
    vinculados_por_b1 = 0
    updates = []

    for row in rows:
        raw_sku = row[col["No Parte"]].value
        sku = str(raw_sku).strip() if raw_sku else ""
        try:
            existencia = int(row[col["Existencia"]].value)
        except (TypeError, ValueError):
            existencia = None

        info = None
        nota_vinculo = None

        # El codigo B1 es la llave universal (100% del conteo, sin duplicados),
        # asi que sirve tanto de respaldo cuando falta el SKU como de rescate
        # cuando el SKU de Shopify no es el del POS.
        raw_b1 = row[col["Codigo B1"]].value if "Codigo B1" in col else None
        b1 = str(raw_b1).strip() if raw_b1 not in (None, "") else ""
        info_b1 = barcode_map.get(b1) if b1 else None

        if sku and sku_counts.get(sku, 0) > 1:
            estatus, nota, fill = "gris", f"Código duplicado en el conteo ({sku_counts[sku]} veces)", COLOR_GRAY
        elif sku and sku in sku_map:
            info = sku_map[sku]
        elif info_b1 is not None:
            info = info_b1
            nota_vinculo = f"Vinculado por código B1 {b1}"
            vinculados_por_b1 += 1
        elif sku:
            estatus, nota, fill = "rojo", "No existe en Shopify", COLOR_RED
        else:
            descripcion = row[col["Descripción"]].value if "Descripción" in col else None
            info, nota_vinculo = find_by_name(descripcion, title_index)
            if info is None:
                estatus, nota, fill = "gris", nota_vinculo, COLOR_GRAY
            vinculados_por_nombre += 1 if info else 0

        if info is not None:
            antes = info["available"]
            row[col_antes].value = antes
            id_log = sku or info["product_title"]

            if existencia is None:
                estatus, nota, fill = "gris", "Existencia no numérica", COLOR_GRAY
            elif existencia < 0:
                estatus = "rojo"
                nota = f"Conteo negativo ({existencia}) — tratado como agotado, revisar POS"
                fill = COLOR_RED
                if antes != 0:
                    updates.append((id_log, info, 0))
            elif existencia == 0:
                estatus, fill = "rojo", COLOR_RED
                nota = "Agotado"
                if antes != 0:
                    updates.append((id_log, info, 0))
                else:
                    nota = "Agotado (ya estaba en 0)"
            elif existencia == antes:
                estatus, nota, fill = "verde", "Sin cambios", COLOR_GREEN
            else:
                estatus = "amarillo"
                nota = f"Actualizado de {antes} a {existencia}"
                fill = COLOR_YELLOW
                updates.append((id_log, info, existencia))

            if nota_vinculo:
                nota = f"{nota_vinculo} — {nota}"

        row[col_estatus].value = estatus
        row[col_nota].value = nota
        counts[estatus] += 1
        for cell in row:
            cell.fill = fill

    print(f"Clasificación: {counts}")
    print(f"  (de ellos, {vinculados_por_b1} vinculados por código B1 y "
          f"{vinculados_por_nombre} por nombre)")
    print(f"Actualizando {len(updates)} variantes en Shopify...")
    errors = 0
    for i, (id_log, info, new_qty) in enumerate(updates):
        result = set_inventory(token, store, location_id, info["inventory_item_id"], new_qty)
        if "error" in result:
            print(f"  ERROR {id_log}: {result['error']}", file=sys.stderr)
            errors += 1
        time.sleep(0.5)
        if (i + 1) % 20 == 0:
            print(f"  {i + 1}/{len(updates)}")

    summary_ws = wb.create_sheet("Resumen")
    summary_ws.append(["Estatus", "Cantidad"])
    for k, v in counts.items():
        summary_ws.append([k, v])
    summary_ws.append(["Actualizaciones aplicadas", len(updates) - errors])
    summary_ws.append(["Errores al actualizar", errors])
    summary_ws.append(["Vinculados por código B1", vinculados_por_b1])
    summary_ws.append(["Vinculados por nombre (último recurso)", vinculados_por_nombre])

    wb.save(output_path)
    print(f"Listo. {len(updates) - errors} actualizaciones aplicadas, {errors} errores.")
    print(f"Excel de resultado: {output_path}")


if __name__ == "__main__":
    main()
