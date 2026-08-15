#!/usr/bin/env python3
"""Guarda el "Codigo B1" del POS en el campo `barcode` de cada producto de Shopify.

Por que existe
--------------
El conteo fisico y Shopify se cruzan por SKU (`No Parte` <-> `variant.sku`),
pero ese cruce deja huecos: hay filas del POS sin `No Parte`, y hay
productos cuyo SKU en Shopify viene del catalogo del fabricante
(`632252557`) en vez del codigo interno del POS (`15ANZUEL658EC`). El 15
de agosto de 2026 eso dejaba 17 productos de Shopify que el conteo nunca
tocaba.

El export del POS si trae un identificador universal: `Codigo B1`, presente
en el 100% de las filas y sin duplicados. Guardandolo en `barcode` —un
campo que esta completamente vacio en esta tienda— el cruce pasa a ser
exacto por SKU **o** por codigo B1, sin heuristicas de texto.

Este script llena ese campo para los productos que ya casan por SKU. Los
que no casan se reportan al final para revisarlos a mano: mapearlos
automaticamente por parecido de nombre produce falsos positivos peligrosos
(p. ej. "Pistola Glock 19" contra "Pistola P1911 Swiss Arms", que son
armas distintas).

Uso:
    SHOPIFY_ADMIN_TOKEN=shpat_... python3 scripts/vincular-codigo-b1.py conteo.xlsx --dry-run
    SHOPIFY_ADMIN_TOKEN=shpat_... python3 scripts/vincular-codigo-b1.py conteo.xlsx

El Excel debe traer las columnas `Codigo B1`, `No Parte` y `Descripción`.

Variables de entorno:
    SHOPIFY_ADMIN_TOKEN  (obligatoria) token Admin API con write_products
    SHOPIFY_STORE        dominio myshopify (default: wfuxvx-yn.myshopify.com)
"""

import argparse
import json
import os
import sys
import time
import urllib.error
import urllib.request

import openpyxl

API_VERSION = "2024-01"


def api_request(method, path, token, store, body=None):
    url = f"https://{store}/admin/api/{API_VERSION}{path}"
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("X-Shopify-Access-Token", token)
    if data:
        req.add_header("Content-Type", "application/json")
    for intento in range(5):
        try:
            with urllib.request.urlopen(req) as resp:
                return json.loads(resp.read()), resp.headers.get("Link", "")
        except urllib.error.HTTPError as e:
            if e.code == 429:
                time.sleep(2 * (intento + 1))
                continue
            return {"error": f"HTTP {e.code}: {e.read().decode()[:300]}"}, ""
        except (urllib.error.URLError, TimeoutError, OSError) as e:
            print(f"  red inestable ({type(e).__name__}), reintento {intento + 1}/5...", file=sys.stderr)
            time.sleep(2 * (intento + 1))
    return {"error": "se agotaron los reintentos"}, ""


def next_page_url(link_header):
    for part in (link_header or "").split(","):
        if 'rel="next"' in part:
            return part.split(";")[0].strip().strip("<>")
    return None


def fetch_all_products(token, store):
    productos = []
    url = f"https://{store}/admin/api/{API_VERSION}/products.json?limit=250"
    while url:
        req = urllib.request.Request(url)
        req.add_header("X-Shopify-Access-Token", token)
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
            productos.extend(data.get("products", []))
            url = next_page_url(resp.headers.get("Link", ""))
        time.sleep(0.5)
    return productos


def leer_conteo(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {n: i for i, n in enumerate(headers)}
    for requerida in ("Codigo B1", "No Parte", "Descripción"):
        if requerida not in col:
            print(f"El Excel debe traer la columna '{requerida}'", file=sys.stderr)
            sys.exit(1)

    por_sku, sin_sku = {}, []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        b1 = fila[col["Codigo B1"]]
        if b1 in (None, ""):
            continue
        b1 = str(b1).strip()
        sku = fila[col["No Parte"]]
        desc = fila[col["Descripción"]]
        if sku in (None, ""):
            sin_sku.append({"b1": b1, "desc": desc})
        else:
            por_sku[str(sku).strip()] = {"b1": b1, "desc": desc}
    return por_sku, sin_sku


def actualizar_barcode(token, store, variant_id, barcode):
    body = {"variant": {"id": variant_id, "barcode": barcode}}
    data, _ = api_request("PUT", f"/variants/{variant_id}.json", token, store, body)
    return data


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("excel", help="Export del conteo fisico con la columna 'Codigo B1'")
    parser.add_argument(
        "--dry-run", action="store_true", help="Muestra que haria, sin escribir en Shopify"
    )
    args = parser.parse_args()

    token = os.environ.get("SHOPIFY_ADMIN_TOKEN")
    if not token:
        print("Falta SHOPIFY_ADMIN_TOKEN", file=sys.stderr)
        sys.exit(1)
    store = os.environ.get("SHOPIFY_STORE", "wfuxvx-yn.myshopify.com")

    por_sku, sin_sku = leer_conteo(args.excel)
    print(f"Conteo: {len(por_sku)} filas con No Parte, {len(sin_sku)} sin el")

    print("Descargando productos de Shopify...")
    productos = fetch_all_products(token, store)
    print(f"  {len(productos)} productos")

    pendientes, ya_ok, sin_match = [], 0, []
    for producto in productos:
        variantes = producto.get("variants", [])
        if len(variantes) != 1:
            # Con mas de una variante no se sabe a cual corresponde el conteo.
            sin_match.append((producto["title"], "tiene mas de 1 variante"))
            continue
        variante = variantes[0]
        sku = (variante.get("sku") or "").strip()
        fila = por_sku.get(sku)
        if not fila:
            sin_match.append((producto["title"], f"SKU '{sku}' no aparece en el conteo"))
            continue
        actual = (variante.get("barcode") or "").strip()
        if actual == fila["b1"]:
            ya_ok += 1
        else:
            pendientes.append((producto["title"], variante["id"], fila["b1"], actual))

    print()
    print(f"  ya tienen el codigo B1 correcto : {ya_ok}")
    print(f"  por escribir                    : {len(pendientes)}")
    print(f"  sin correspondencia en el conteo: {len(sin_match)}  <- revisar a mano")

    if args.dry_run:
        print("\n--dry-run: no se escribio nada. Ejemplos:")
        for titulo, _, b1, actual in pendientes[:10]:
            previo = f" (tenia '{actual}')" if actual else ""
            print(f"  {b1:<15} {titulo[:55]}{previo}")
    elif pendientes:
        print()
        errores = 0
        for i, (titulo, variant_id, b1, _) in enumerate(pendientes):
            resultado = actualizar_barcode(token, store, variant_id, b1)
            if "error" in resultado:
                print(f"  ERROR {titulo[:45]}: {resultado['error']}", file=sys.stderr)
                errores += 1
            time.sleep(0.5)
            if (i + 1) % 50 == 0:
                print(f"  {i + 1}/{len(pendientes)}")
        print(f"\nEscritos: {len(pendientes) - errores}  ({errores} errores)")

    if sin_match:
        print("\n--- Productos de Shopify SIN codigo B1 (revisar a mano) ---")
        for titulo, motivo in sin_match:
            print(f"  {titulo[:60]:<60} {motivo}")
        print(
            "\nPara cada uno: ubica su fila en el conteo fisico, copia su 'Codigo B1'\n"
            "y pegalo en el campo 'Codigo de barras' del producto en Shopify.\n"
            "NO se mapean automaticamente por parecido de nombre: productos que solo\n"
            "difieren en calibre o talla se confunden entre si."
        )


if __name__ == "__main__":
    main()
